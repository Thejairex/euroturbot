from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from core.exceptions import ExcelError


class ExcelReader:
    def __init__(self, filepath: str | Path):
        self.filepath = Path(filepath)
        if not self.filepath.exists():
            raise ExcelError(f"Archivo no encontrado: {self.filepath}")

    def read_all(self, sheet_name: str | int = 0) -> list[dict[str, Any]]:
        df = pd.read_excel(self.filepath, sheet_name=sheet_name, dtype=str)
        return df.where(pd.notna(df), None).to_dict(orient="records")

    def read_rows(self, sheet_name: str | int = 0) -> list[list[Any]]:
        wb = load_workbook(self.filepath, data_only=True)
        ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]
        return [[cell.value for cell in row] for row in ws.iter_rows()]

    def find_row(self, column: str, value: str, sheet_name: str | int = 0) -> dict[str, Any] | None:
        data = self.read_all(sheet_name)
        for row in data:
            if str(row.get(column, "")).strip().lower() == value.strip().lower():
                return row
        return None

    def find_all(self, column: str, value: str, sheet_name: str | int = 0) -> list[dict[str, Any]]:
        data = self.read_all(sheet_name)
        return [
            row for row in data
            if str(row.get(column, "")).strip().lower() == value.strip().lower()
        ]

    def sheet_names(self) -> list[str]:
        wb = load_workbook(self.filepath, read_only=True)
        return wb.sheetnames
