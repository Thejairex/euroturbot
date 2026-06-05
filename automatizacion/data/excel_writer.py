from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from core.exceptions import ExcelError


class ExcelWriter:
    def __init__(self, filepath: str | Path):
        self.filepath = Path(filepath)
        self.filepath.parent.mkdir(parents=True, exist_ok=True)

    def write_rows(self, data: list[dict[str, Any]], sheet_name: str = "Resultados") -> str:
        df = pd.DataFrame(data)
        df.to_excel(self.filepath, sheet_name=sheet_name, index=False)
        return str(self.filepath)

    def append_rows(self, data: list[dict[str, Any]], sheet_name: str = "Resultados") -> str:
        if not self.filepath.exists():
            return self.write_rows(data, sheet_name)

        existing = pd.read_excel(self.filepath, sheet_name=sheet_name)
        new_df = pd.DataFrame(data)
        combined = pd.concat([existing, new_df], ignore_index=True)
        combined.to_excel(self.filepath, sheet_name=sheet_name, index=False)
        return str(self.filepath)

    def update_cell(self, row: int, col: int, value: Any, sheet_name: str = "Resultados") -> None:
        wb = load_workbook(self.filepath)
        ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.active
        ws.cell(row=row, column=col, value=value)
        wb.save(self.filepath)
